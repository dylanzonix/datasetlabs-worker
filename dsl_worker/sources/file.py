"""file source — parse uploaded CSV/XLSX or files written by code_exec.

CSV/XLSX get auto-detected and parsed; rows = file rows, columns inferred
from headers (with light type sniffing). Other file types return an error
with a hint to use `code_exec` to transform first.

Predictable in the sense that headers map directly to columns — no agent-side
mapping step needed. (Agent can still rename/retype columns later via
`column_map_set` if desired.)
"""

from __future__ import annotations

import csv
import io
import logging
import os
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

from dsl_worker.sources.base import FetchResult, SourceAdapter, SourceDescription, register


log = logging.getLogger(__name__)


_INT_RE = re.compile(r"^-?\d+$")
_FLOAT_RE = re.compile(r"^-?\d+\.\d+$")
_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}")
_UUID_RE = re.compile(r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$", re.I)


def _read_project_file_bytes(
    file_id: str, project_id: str,
) -> tuple[Optional[bytes], Optional[str]]:
    """Read an uploaded project file's raw bytes from Azure Blob.

    Returns (bytes, filename) or (None, None) if not found.
    """
    if not _UUID_RE.match(file_id):
        return None, None
    try:
        from dsl_api.db import SessionLocal
        from sqlalchemy import text as sa_text
        db = SessionLocal()
        try:
            row = db.execute(
                sa_text(
                    "SELECT blob_path, filename FROM project_files "
                    "WHERE id=:fid AND deleted_at IS NULL AND status='uploaded'"
                ),
                {"fid": file_id},
            ).fetchone()
        finally:
            db.close()
        if row and row[0]:
            from dsl_api.azure.blob import get_blob_client
            blob_client = get_blob_client(row[0])
            return blob_client.download_blob().readall(), row[1]
    except Exception as e:
        log.info("_read_project_file_bytes: %s failed: %s", file_id, e)
    return None, None
_URL_RE = re.compile(r"^https?://")
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _sniff_type(samples: List[str]) -> str:
    """Infer a column's type from a handful of sample values. Conservative."""
    non_empty = [s for s in samples if s]
    if not non_empty:
        return "text"
    if all(_URL_RE.match(s) for s in non_empty):
        return "url"
    if all(_EMAIL_RE.match(s) for s in non_empty):
        return "email"
    if all(_DATE_RE.match(s) for s in non_empty):
        return "date"
    if all(_FLOAT_RE.match(s) or _INT_RE.match(s) for s in non_empty):
        return "number"
    if all(s.lower() in {"true", "false", "yes", "no", "1", "0"} for s in non_empty):
        return "bool"
    return "text"


class FileAdapter(SourceAdapter):
    name = "file"
    label = "Uploaded files"
    favicon_url = None  # frontend renders a generic file icon
    predictable = True  # CSV headers ARE the schema

    def describe(
        self,
        query_params: Dict[str, Any],
        source: Optional[str] = None,
    ) -> SourceDescription:
        qp = query_params or {}
        # `file_id` may be a single id or a list when multiple files are
        # processed as one source. Filenames are resolved by chat before
        # this is rendered; here we just show the raw ids if filenames
        # aren't attached.
        files = qp.get("files") or qp.get("file_ids") or qp.get("file_id")
        if isinstance(files, str):
            files_list: List[str] = [files]
        elif isinstance(files, list):
            files_list = [str(f) for f in files]
        else:
            files_list = []
        n = len(files_list)
        headline = (
            "1 attached file" if n == 1
            else f"{n} attached files" if n > 1
            else "Attached files"
        )
        # `query` here is the iteration semantic — same query applies to all
        # attached files. This is the field the user wrote in chat / panel.
        if qp.get("query"):
            headline = f"{headline} — {qp['query']}"
        detail_lines = []
        if files_list:
            detail_lines.append("**Files:** " + ", ".join(files_list))
        if qp.get("description"):
            detail_lines.append(qp["description"])
        return SourceDescription(
            kind=self.name,
            label=self.label,
            query_text=headline,
            details="\n\n".join(detail_lines),
            favicon_url=self.favicon_url,
        )

    def validate_query_params(self, query_params: Dict[str, Any]) -> Optional[str]:
        if "file_id" not in query_params:
            return "file source requires `file_id`"
        return None

    async def fetch(
        self,
        query_params: Dict[str, Any],
        n: int,
        prior_cursor: Optional[Dict[str, Any]] = None,
    ) -> FetchResult:
        file_id = query_params["file_id"]
        path = Path(file_id)
        # 1) Already a real disk path → use it.
        # 2) Looks like a UUID → it's a project_files row; read from Azure Blob.
        # 3) Otherwise fall back to the candidate store (code_exec output).
        if not path.exists():
            project_id = query_params.get("_project_id")
            blob_bytes: Optional[bytes] = None
            blob_filename: Optional[str] = None

            blob_bytes, blob_filename = _read_project_file_bytes(file_id, project_id or "")
            if blob_bytes is None:
                pass  # fall through to candidates lookup

            if blob_bytes is None and project_id:
                try:
                    from dsl_worker.chat.candidates import read_candidates_bytes
                    blob_bytes = read_candidates_bytes(project_id, file_id)
                    blob_filename = file_id
                except Exception as e:
                    log.info("file source: candidate lookup for %s failed: %s", file_id, e)

            if blob_bytes is None:
                return FetchResult(
                    rows=[], schema=[], cost_credits=0.0, exhausted=True,
                )

            import tempfile
            suffix = Path(blob_filename or file_id).suffix or ".csv"
            tmp = tempfile.NamedTemporaryFile(
                delete=False, suffix=suffix, mode="wb",
            )
            tmp.write(blob_bytes)
            tmp.close()
            path = Path(tmp.name)

        rows: List[Dict[str, Any]] = []
        if path.suffix.lower() == ".csv":
            with path.open("r", encoding="utf-8", errors="replace") as f:
                reader = csv.DictReader(f)
                offset = int((prior_cursor or {}).get("offset", 0))
                for i, row in enumerate(reader):
                    if i < offset:
                        continue
                    if len(rows) >= n:
                        break
                    rows.append({k: v for k, v in row.items() if k})
        else:
            # XLSX support — try openpyxl if installed. Otherwise error.
            try:
                import openpyxl
                wb = openpyxl.load_workbook(path, read_only=True)
                ws = wb.active
                headers = [str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
                offset = int((prior_cursor or {}).get("offset", 0))
                for i, row_cells in enumerate(ws.iter_rows(min_row=2)):
                    if i < offset:
                        continue
                    if len(rows) >= n:
                        break
                    rows.append({h: (c.value if c.value is not None else "") for h, c in zip(headers, row_cells)})
            except ImportError:
                return FetchResult(
                    rows=[],
                    schema=[],
                    cost_credits=0.0,
                    exhausted=True,
                )
            except Exception as e:
                log.exception("file adapter xlsx parse failed: %s", e)
                return FetchResult(rows=[], schema=[], cost_credits=0.0, exhausted=True)

        # Sniff types from first 20 rows to populate default_columns dynamically.
        if rows:
            schema_keys = list({k for r in rows for k in r.keys()})
            sniffed: List[Dict[str, str]] = []
            for k in schema_keys:
                samples = [str(r.get(k, "") or "") for r in rows[:20]]
                sniffed.append({"source_field": k, "column_name": k, "type": _sniff_type(samples)})
            self.default_columns = sniffed

        return FetchResult(
            rows=rows,
            schema=list(rows[0].keys()) if rows else [],
            cost_credits=0.0,
            exhausted=len(rows) < n,
            cursor={"offset": int((prior_cursor or {}).get("offset", 0)) + len(rows)},
        )


register(FileAdapter())
