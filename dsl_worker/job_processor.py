"""
Job Processor — orchestrator-driven pipeline.
"""

import asyncio
import json
import logging
import os
import shutil
import socket
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional
from uuid import UUID

from openai import AsyncOpenAI
from sqlalchemy.orm import Session
from azure.storage.blob import BlobServiceClient, generate_blob_sas, BlobSasPermissions

from dsl_api.models.project import Project
from dsl_api.models.project_version import ProjectVersion
from dsl_api.models.project_event import ProjectEvent
from dsl_api.models.chat_message import ChatMessage
from dsl_api.models.project_file import ProjectFile
from dsl_api.models.project_connector import ProjectConnector
from dsl_api.models.sample import Sample

from dsl_worker.config import settings

# Langfuse is optional
try:
    from langfuse import get_client as _get_langfuse_client, propagate_attributes

    def _get_langfuse():
        try:
            return _get_langfuse_client()
        except Exception:
            return None
except ImportError:
    def _get_langfuse():
        return None

    def propagate_attributes(**kwargs):
        from contextlib import contextmanager
        @contextmanager
        def _noop():
            yield
        return _noop()
from dsl_worker.project_state import ProjectState
from dsl_worker.billing import CostTracker, TrackedOpenAIClient
from dsl_worker.checkpoint import CheckpointManager

from dsl_worker.agents import OrchestratorAgent
from dsl_worker.agents.row import DedupStore
from dsl_worker.infra.generation_pool import GenerationWorkerPool
from sandbox_service import SandboxClient

logger = logging.getLogger(__name__)

WORKER_ID = f"{socket.gethostname()}-{os.getpid()}"


class JobProcessor:
    """
    Job processor for the V6 pipeline.

    Flow: orchestrator (dispatches subagents internally) → row generators (parallel)
    The orchestrator stays in the loop, spawning yielders/synthesizers and submitting
    seeds. Generation runs in the background as seeds flow in.
    Checkpoints saved to Azure Blob for pause/resume.
    """

    def __init__(
        self,
        db_session_factory,
        openai_client: AsyncOpenAI,
        blob_service_client: BlobServiceClient,
    ):
        self.SessionLocal = db_session_factory
        self.raw_openai_client = openai_client
        self.blob_service_client = blob_service_client
        self.should_stop = False

        # Sandbox client (shared HTTP connection — each agent creates its own session)
        self._sandbox: Optional[SandboxClient] = None
        self._workspace_dir: Optional[Path] = None
        # MCP tools for current job (set during _run_pipeline)
        self._mcp_tools: list = []

    def request_stop(self):
        """Request graceful stop."""
        logger.warning("[Worker] Stop requested")
        self.should_stop = True

    def _make_stop_checker(self, state: ProjectState, stop_event: asyncio.Event):
        """Create a stop checker that refreshes state before checking.

        Also sets stop_event so all agents wake up instantly once stop is detected,
        rather than each waiting for their own poll cycle.
        """
        def checker():
            state.refresh()
            should_stop = self.should_stop or state.paused or stop_event.is_set()
            if should_stop and not stop_event.is_set():
                stop_event.set()
            return should_stop
        return checker

    async def process_job(self, message_body: Dict[str, Any]) -> bool:
        """Process a job."""
        # Reset stop flag from previous job (credit exhaustion sets this)
        self.should_stop = False

        project_id_str = message_body.get("project_id")
        version_id_str = message_body.get("version_id")

        if not project_id_str or not version_id_str:
            logger.error("Invalid message: missing IDs")
            return False

        project_id = UUID(project_id_str)
        version_id = UUID(version_id_str)

        logger.info(f"Starting job: project={project_id}, version={version_id}")

        db: Session = self.SessionLocal()

        try:
            # Validate
            project = db.query(Project).filter(Project.id == project_id).first()
            if not project:
                logger.error(f"Project not found: {project_id}")
                return False

            version = db.query(ProjectVersion).filter(ProjectVersion.id == version_id).first()
            if not version:
                logger.error(f"Version not found: {version_id}")
                return False

            # Check stale
            if project.current_version_id != version_id:
                logger.warning("Stale message")
                return True

            # Check status
            if version.status not in ("running", "pause_requested"):
                logger.warning(f"Version not running: {version.status}")
                return True

            if version.status == "pause_requested":
                self._handle_pause(db, project, version, None, "Pause before start")
                return True

            # Set started
            if version.started_at is None:
                version.started_at = datetime.now(timezone.utc)
                db.commit()

            logger.info(f"Project: {project.name}, Version: {version.version_number}")
            logger.info(f"Target: {version.num_samples} samples")

            # Create tracked client
            tracked_client = TrackedOpenAIClient(self.raw_openai_client)

            # Create cost tracker
            cost_tracker = CostTracker(
                db=db,
                user_id=project.user_id,
                project_id=project_id,
                compute_cost_per_credit=settings.compute_cost_per_credit,
                charge_threshold_cents=settings.billing_charge_threshold_cents,
                charge_interval_seconds=settings.billing_charge_interval_seconds,
            )

            # Check balance
            can_continue, stop_reason = cost_tracker.check_balance_and_charge()
            if not can_continue:
                self._handle_force_stop(db, project, version, cost_tracker, stop_reason)
                return False

            # Initialize state
            state = ProjectState(db, project_id, version_id)

            # Emit running event
            self._emit_event(db, project, version, "running", "Worker started")

            # Run pipeline wrapped in Langfuse trace.
            # propagate_attributes sets session_id/user_id/tags on every
            # observation created within the block, including sub-agent spans
            # in asyncio tasks that inherit the contextvars context.
            # Select pipeline version
            pipeline_fn = self._run_pipeline
            if settings.pipeline_version == "v13":
                pipeline_fn = self._run_pipeline_v13
                logger.info("[Pipeline] Using V13 pipeline")

            with propagate_attributes(
                session_id=str(project_id),
                user_id=str(project.user_id),
                tags=[project.name, f"v{version.version_number}"],
                metadata={
                    "project_id": str(project_id),
                    "version_id": str(version_id),
                    "num_samples": str(version.num_samples),
                },
            ):
                langfuse = _get_langfuse()
                if langfuse:
                    with langfuse.start_as_current_observation(
                        as_type="span",
                        name=f"job:{project.name} v{version.version_number}",
                    ):
                        result = await pipeline_fn(
                            db, project, version, state, tracked_client, cost_tracker,
                        )
                else:
                    result = await pipeline_fn(
                        db, project, version, state, tracked_client, cost_tracker
                    )
            return result

        except Exception as e:
            logger.exception(f"Job error: {e}")

            try:
                version = db.query(ProjectVersion).filter(ProjectVersion.id == version_id).first()
                project = db.query(Project).filter(Project.id == project_id).first()
                if version:
                    version.status = "failed"
                    version.error = str(e)
                    version.finished_at = datetime.now(timezone.utc)
                    if project:
                        self._emit_event(db, project, version, "failed", "Error", {"error": str(e)})
                    db.commit()
            except Exception as db_err:
                logger.error(f"Failed to update status: {db_err}")

            return False

        finally:
            # Flush Langfuse traces before cleanup
            langfuse = _get_langfuse()
            if langfuse:
                langfuse.flush()
            await self._cleanup()
            db.close()

    async def _run_pipeline(
        self,
        db: Session,
        project: Project,
        version: ProjectVersion,
        state: ProjectState,
        tracked_client: TrackedOpenAIClient,
        cost_tracker: CostTracker,
    ) -> bool:
        """Run the V5 pipeline: orchestrator → seed yielders → row generators."""

        # Create workspace (only downloads/ needed locally — uploads go direct to sandbox)
        workspace_dir = Path(f"./workspace_{project.id}_{version.id}")
        workspace_dir.mkdir(parents=True, exist_ok=True)
        (workspace_dir / "downloads").mkdir(exist_ok=True)
        self._workspace_dir = workspace_dir

        # Restore browser downloads from blob (survives pause/resume)
        self._restore_downloads_from_blob(project.id, workspace_dir)

        # Generate SAS URLs for uploaded files (sandbox service fetches them directly)
        uploaded_file_urls = self._generate_file_urls(db, project.id)

        # Initialize sandbox client
        self._sandbox = SandboxClient(settings.sandbox_service_url, timeout=150.0)
        await self._sandbox.__aenter__()

        # Initialize checkpoint manager
        checkpoint_mgr = CheckpointManager(
            blob_service_client=self.blob_service_client,
            container_name=settings.azure_storage_container_name,
            project_id=project.id,
            version_id=version.id,
        )
        checkpoint = await checkpoint_mgr.initialize()

        stop_event = asyncio.Event()
        stop_checker = self._make_stop_checker(state, stop_event)

        # Track whether we stopped due to credit exhaustion
        credit_exhausted = False

        # Per-turn cost callback — fires on every API call in every agent.
        async def on_cost(cost_usd: float, label: str):
            nonlocal credit_exhausted
            try:
                cost_tracker.add_cost(phase=label, cost_usd=cost_usd)
                cost_tracker.charge_if_needed()
                await checkpoint_mgr.add_cost(cost_usd)

                # Stop the pipeline if credits are exhausted
                if not credit_exhausted and not cost_tracker.has_sufficient_balance():
                    credit_exhausted = True
                    logger.warning("[Billing] Credits exhausted — stopping pipeline")
                    self.should_stop = True
                    stop_event.set()
            except Exception as e:
                logger.error(f"[Billing] on_cost callback error: {e}")

        # Seed cost tracker from checkpoint if resuming
        if checkpoint.total_cost_usd > 0:
            cost_tracker.seed_from_checkpoint(checkpoint.total_cost_usd)

        # --- Detect resume: check if rows already exist for this version ---
        db.refresh(version)
        existing_row_count = version.generated_count or 0
        is_resume = existing_row_count > 0

        if is_resume:
            logger.info(
                f"[Pipeline] RESUMING: {existing_row_count}/{state.num_samples} rows "
                f"already generated, checkpoint cost=${checkpoint.total_cost_usd:.4f}"
            )

        # Load conversation history
        chat_history = self._load_chat_history(db, project.id)

        # Build uploaded files metadata for orchestrator context
        project_files = (
            db.query(ProjectFile)
            .filter(
                ProjectFile.project_id == project.id,
                ProjectFile.deleted_at.is_(None),
                ProjectFile.status == "uploaded",
            )
            .all()
        )
        uploaded_files = [
            {
                "filename": f.filename,
                "content_type": f.content_type,
                "size_bytes": f.size_bytes,
            }
            for f in project_files
        ]

        # Enrich file metadata — inspect CSVs/JSONs for columns, row counts, previews
        if uploaded_files and uploaded_file_urls:
            uploaded_files = await self._enrich_file_metadata(
                uploaded_files, uploaded_file_urls,
            )

        # Load connectors for MCP tool injection
        connectors = (
            db.query(ProjectConnector)
            .filter(
                ProjectConnector.project_id == project.id,
                ProjectConnector.deleted_at.is_(None),
            )
            .all()
        )
        mcp_tools = self._build_mcp_tools(connectors)
        self._mcp_tools = mcp_tools

        logger.info(f"[Pipeline] Starting V11 pipeline with {len(chat_history)} chat messages")

        # === Shared state ===
        # Seed from DB so resume knows how many rows already exist
        generation_stats = {
            "rows_generated": existing_row_count,
            "errors": 0,
            "skipped": 0,
            "in_progress": 0,
            "total_cost": checkpoint.total_cost_usd,
        }

        # Feedback context for re-planning iterations
        feedback_context = None

        # Check if this is a feedback iteration (version has previous config in recipe)
        if checkpoint.recipe:
            try:
                prev_config = json.loads(checkpoint.recipe)
                # Check for user feedback messages since version started
                feedback_msg = (
                    db.query(ChatMessage)
                    .filter(
                        ChatMessage.project_id == project.id,
                        ChatMessage.role == "user",
                        ChatMessage.created_at > version.started_at,
                    )
                    .order_by(ChatMessage.created_at.desc())
                    .first()
                )
                if feedback_msg:
                    feedback_context = {
                        "previous_config": prev_config,
                        "user_feedback": feedback_msg.content,
                    }
            except (json.JSONDecodeError, Exception):
                pass

        # --- Progress tracking + checkpointing ---
        progress_counters: Dict[str, Any] = {"phase": "orchestrating", "steps": []}
        last_progress_flush = time.time()
        last_langfuse_flush = time.time()

        # User-friendly step labels for orchestrator tool calls
        STEP_LABELS: Dict[str, str] = {
            "code_exec": "Setting up...",
            "create_harvester": "Finding sources...",
            "apollo_search": "Searching database...",
            "apollo_search_companies": "Searching database...",
            "process": "Filling in details...",
        }

        def on_tool_call(agent_label: str, tool_name: str):
            nonlocal last_progress_flush, last_langfuse_flush

            # Only orchestrator tool calls become steps
            if agent_label == "orchestrator" and tool_name in STEP_LABELS:
                steps = progress_counters.get("steps", [])
                label = STEP_LABELS[tool_name]
                # Mark previous step as done
                if steps and not steps[-1].get("done"):
                    steps[-1]["done"] = True
                # Don't duplicate consecutive identical labels
                if not steps or steps[-1].get("label") != label:
                    steps.append({"label": label, "done": False})
                progress_counters["steps"] = steps

            phase_map = {
                "code_exec": "researching",
                "create_harvester": "harvesting",
                "apollo_search": "harvesting",
                "apollo_search_companies": "harvesting",
                "process": "generating",
                "close_source": "generating",
            }
            if tool_name in phase_map:
                progress_counters["phase"] = phase_map[tool_name]

            now = time.time()
            if now - last_progress_flush >= 2.0:
                merged = {**progress_counters, **generation_stats}
                version.progress_detail = merged
                db.commit()
                last_progress_flush = now
            if now - last_langfuse_flush >= 10.0:
                lf = _get_langfuse()
                if lf:
                    lf.flush()
                last_langfuse_flush = now


        # --- Browser session tracking ---
        browser_sessions: Dict[str, Dict[str, str]] = {}

        def on_browser_started(live_url: str, session_id: str):
            browser_sessions[session_id] = {
                "live_url": live_url,
                "session_id": session_id,
            }
            progress_counters["browser_sessions"] = list(browser_sessions.values())
            logger.info(f"[Pipeline] Cloud browser started: {live_url}")

        def on_browser_stopped(session_id: str):
            browser_sessions.pop(session_id, None)
            progress_counters["browser_sessions"] = list(browser_sessions.values())
            logger.info(f"[Pipeline] Cloud browser stopped: {session_id}")

        version.progress_detail = {"phase": "orchestrating"}
        db.commit()

        # ====================================================================
        # V11: Orchestrator-driven batch pipeline
        # ====================================================================

        dedup_store = DedupStore()

        # Seed DedupStore with existing rows so resume doesn't create duplicates
        if is_resume:
            from dsl_worker.agents.row import _tokenize
            existing_samples = (
                db.query(Sample)
                .filter(Sample.version_id == version.id)
                .all()
            )
            for sample in existing_samples:
                row_data = sample.row or {}
                row_id = str(sample.id)
                dedup_store._submitted[row_id] = row_data
                for col_name, value in row_data.items():
                    dedup_store._token_cache[(row_id, col_name)] = _tokenize(value)
            logger.info(f"[Pipeline] Seeded DedupStore with {len(existing_samples)} existing rows")

        from dsl_worker.infra.bu_client import BUClient
        bu_client = BUClient(
            api_key=settings.browser_use_api_key,
            model="bu-mini",
            stop_event=stop_event,
        )

        # Apollo.io client (optional — for B2B contact/company data)
        apollo_client = None
        if settings.apollo_api_key:
            from dsl_worker.infra.apollo_client import ApolloClient
            apollo_client = ApolloClient(api_key=settings.apollo_api_key)
            logger.info("[Pipeline] Apollo.io client initialized")

        # Google Maps Places API (optional — for local business data)
        google_maps_client = None
        if settings.google_api_key:
            from dsl_worker.infra.google_maps_client import GoogleMapsClient
            google_maps_client = GoogleMapsClient(api_key=settings.google_api_key)
            logger.info("[Pipeline] Google Maps client initialized")

        # YouTube Data API (optional — for video/channel data)
        youtube_client = None
        if settings.google_api_key:
            from dsl_worker.infra.youtube_client import YouTubeClient
            youtube_client = YouTubeClient(api_key=settings.google_api_key)
            logger.info("[Pipeline] YouTube client initialized")

        # Row saver — reuses GenerationWorkerPool's _save_row for DB writes
        row_saver = GenerationWorkerPool(
            workspace_dir=workspace_dir,
            openai_client=tracked_client,
            db_session=db,
            project_id=project.id,
            version_id=version.id,
        )

        async def save_row(row: Dict, tags: Dict = None) -> Optional[str]:
            """Save a row to DB. Returns row_id or None if target reached."""
            db.refresh(version)
            if (version.generated_count or 0) >= state.num_samples:
                logger.info("[Pipeline] Discarding row: target already reached")
                return None
            return await row_saver._save_row(row, tags=tags or {})

        # Build resume context for orchestrator
        resume_context = None
        if is_resume:
            resume_context = {
                "rows_generated": existing_row_count,
                "target": state.num_samples,
                "remaining": state.num_samples - existing_row_count,
                "prior_cost_usd": checkpoint.total_cost_usd,
                "previous_sources": (checkpoint.source_stats or {}).get("sources", []),
            }

        # Checkpoint callback — called by orchestrator after state changes
        def on_checkpoint(orch):
            try:
                state_data = orch.export_state()
                asyncio.create_task(checkpoint_mgr.save_pipeline_state(
                    orchestrator_messages=state_data["orchestrator_conversation"]["messages"],
                    orchestrator_cost=state_data["orchestrator_conversation"]["total_cost"],
                    orchestrator_turns=state_data["orchestrator_conversation"]["total_turns"],
                    sources=state_data["sources"],
                    generation_stats=state_data["generation_stats"],
                    harvester_counter=state_data["harvester_counter"],
                    apollo_counter=state_data["apollo_counter"],
                    research_counter=state_data["research_counter"],
                ))
            except Exception as e:
                logger.warning(f"[Pipeline] Checkpoint callback error: {e}")

        logger.info("[Pipeline] Running V12 orchestrator")

        # V12: Create dispatcher for automatic candidate → row processing
        from dsl_worker.infra.dispatcher import CandidateDispatcher
        from dsl_worker.agents.row import RowGeneratorAgent, GeneratedRow

        generation_semaphore = asyncio.Semaphore(settings.generation_parallel_samples)
        _save_lock = asyncio.Lock()

        async def generate_row_fn(candidate, source_id):
            """Row generation function passed to dispatcher."""
            rows_done = generation_stats.get("rows_generated", 0)
            if rows_done >= state.num_samples:
                return GeneratedRow(success=False, skipped=True, skip_reason="target reached"), 0.0, False

            def _row_stop():
                if stop_checker and stop_checker():
                    return True
                return generation_stats.get("rows_generated", 0) >= state.num_samples

            agent = RowGeneratorAgent(
                openai_client=tracked_client,
                model=settings.generation_model,
                workspace_dir=workspace_dir,
                chat_history=chat_history,
                dedup_store=dedup_store,
                bu_client=bu_client,
                sandbox=self._sandbox,
                stop_checker=_row_stop,
                stop_event=stop_event,
                blob_service_client=self.blob_service_client,
                project_id=project.id,
                uploaded_file_urls=uploaded_file_urls if uploaded_file_urls else None,
                uploaded_files=uploaded_files if uploaded_files else None,
                apollo_client=apollo_client,
                google_maps_client=google_maps_client,
                youtube_client=youtube_client,
                mcp_tools=mcp_tools,
                on_cost=on_cost,
            )
            try:
                # Get source context from the source state description
                source_context = candidate.source_context or ""
                result = await agent.generate(
                    candidate=candidate.values,
                    schema=state.columns,
                    source_context=source_context,
                )

                saved = False
                if result.success and result.row:
                    async with _save_lock:
                        row_id = await save_row(
                            result.row,
                            tags={"sources": result.sources} if result.sources else {},
                        )
                        saved = row_id is not None
                        if saved:
                            generation_stats["rows_generated"] = (
                                generation_stats.get("rows_generated", 0) + 1
                            )
                elif result.skipped:
                    generation_stats["skipped"] = generation_stats.get("skipped", 0) + 1
                else:
                    generation_stats["errors"] = generation_stats.get("errors", 0) + 1

                return result, result.cost_usd, saved

            except asyncio.CancelledError:
                return GeneratedRow(success=False, skipped=True, skip_reason="stopped"), 0.0, False
            except Exception as e:
                logger.error(f"Row generation error: {e}", exc_info=True)
                return GeneratedRow(success=False, error=str(e)), 0.0, False
            finally:
                try:
                    await agent.cleanup()
                except Exception:
                    pass

        dispatcher = CandidateDispatcher(
            generate_row_fn=generate_row_fn,
            semaphore=generation_semaphore,
            generation_stats=generation_stats,
            num_samples=state.num_samples,
            stop_checker=stop_checker,
        )

        orchestrator = OrchestratorAgent(
            chat_history=chat_history,
            columns=state.columns,
            num_samples=state.num_samples,
            openai_client=tracked_client,
            model=settings.research_model,
            workspace_dir=workspace_dir,
            generation_stats=generation_stats,
            dedup_store=dedup_store,
            save_row=save_row,
            dispatcher=dispatcher,
            harvester_model=settings.seed_yielder_model,
            generation_model=settings.generation_model,
            uploaded_files=uploaded_files if uploaded_files else None,
            bu_client=bu_client,
            apollo_client=apollo_client,
            google_maps_client=google_maps_client,
            youtube_client=youtube_client,
            sandbox=self._sandbox,
            stop_checker=stop_checker,
            stop_event=stop_event,
            blob_service_client=self.blob_service_client,
            project_id=project.id,
            on_tool_call=on_tool_call,
            on_cost=on_cost,
            on_checkpoint=on_checkpoint,
            uploaded_file_urls=uploaded_file_urls if uploaded_file_urls else None,
            mcp_tools=mcp_tools,
            feedback_context=feedback_context,
            resume_context=resume_context,
        )

        # Restore full state from checkpoint if available
        if is_resume and checkpoint.has_v11_state:
            orchestrator.restore_state({
                "orchestrator_conversation": checkpoint.orchestrator_conversation,
                "sources": checkpoint.sources,
                "generation_stats": checkpoint.generation_stats,
                "harvester_counter": checkpoint.harvester_counter,
                "apollo_counter": checkpoint.apollo_counter,
                "research_counter": checkpoint.research_counter,
            })

        try:
            result = await orchestrator.run()
            logger.info(
                f"[Pipeline] Orchestrator finished: "
                f"cost=${orchestrator.cost_usd:.4f}, turns={result.turns_taken}"
            )
        finally:
            # Force-save full state before cleanup — this is the final checkpoint
            # that will be used if this is a pause/stop
            try:
                state_data = orchestrator.export_state()
                await checkpoint_mgr.save_pipeline_state(
                    orchestrator_messages=state_data["orchestrator_conversation"]["messages"],
                    orchestrator_cost=state_data["orchestrator_conversation"]["total_cost"],
                    orchestrator_turns=state_data["orchestrator_conversation"]["total_turns"],
                    sources=state_data["sources"],
                    generation_stats=state_data["generation_stats"],
                    harvester_counter=state_data["harvester_counter"],
                    apollo_counter=state_data["apollo_counter"],
                    research_counter=state_data["research_counter"],
                )
                await checkpoint_mgr.force_save()
            except Exception as e:
                logger.warning(f"[Pipeline] Final checkpoint save error: {e}")

            await orchestrator.cleanup()
            # Brief pause to let any orphaned BU polling tasks settle before
            # closing the httpx client (prevents "client has been closed" errors)
            await asyncio.sleep(0.5)
            await bu_client.close()
            if apollo_client:
                await apollo_client.close()

        # ====================================================================
        # COMPLETE — Check results
        # ====================================================================

        await checkpoint_mgr.set_phase("completed")

        state.refresh()
        if state.paused:
            await checkpoint_mgr.force_save()
            self._handle_pause(db, project, version, cost_tracker)
            return True

        can_continue, stop_reason = cost_tracker.check_balance_and_charge()
        if not can_continue:
            await checkpoint_mgr.force_save()
            self._handle_force_stop(db, project, version, cost_tracker, stop_reason)
            return False

        db.refresh(version)
        generated = version.generated_count or 0
        if generated > 0:
            if generated < state.num_samples:
                logger.warning(
                    f"[Pipeline] Completed with {generated}/{state.num_samples} rows"
                )
            await checkpoint_mgr.force_save()
            self._handle_completion(db, project, version, cost_tracker)
            return True

        await checkpoint_mgr.force_save()
        self._handle_force_stop(db, project, version, cost_tracker, "No rows generated")
        return False

    async def _run_pipeline_v13(
        self,
        db: Session,
        project: Project,
        version: ProjectVersion,
        state: ProjectState,
        tracked_client: TrackedOpenAIClient,
        cost_tracker: CostTracker,
    ) -> bool:
        """Run the V13 pipeline: blocking orchestrator, file-based candidate flow."""
        from dsl_worker.agents.orchestrator_v13 import OrchestratorV13
        from dsl_worker.agents.row import RowGeneratorAgent, GeneratedRow

        # Create workspace
        workspace_dir = Path(f"./workspace_{project.id}_{version.id}")
        workspace_dir.mkdir(parents=True, exist_ok=True)
        (workspace_dir / "downloads").mkdir(exist_ok=True)
        self._workspace_dir = workspace_dir

        # Restore browser downloads from blob (survives pause/resume)
        self._restore_downloads_from_blob(project.id, workspace_dir)

        # Generate SAS URLs for uploaded files
        uploaded_file_urls = self._generate_file_urls(db, project.id)

        # Initialize sandbox client
        self._sandbox = SandboxClient(settings.sandbox_service_url, timeout=150.0)
        await self._sandbox.__aenter__()

        # Initialize checkpoint manager
        checkpoint_mgr = CheckpointManager(
            blob_service_client=self.blob_service_client,
            container_name=settings.azure_storage_container_name,
            project_id=project.id,
            version_id=version.id,
        )
        checkpoint = await checkpoint_mgr.initialize()

        stop_event = asyncio.Event()
        stop_checker = self._make_stop_checker(state, stop_event)

        credit_exhausted = False

        async def on_cost(cost_usd: float, label: str):
            nonlocal credit_exhausted
            try:
                cost_tracker.add_cost(phase=label, cost_usd=cost_usd)
                cost_tracker.charge_if_needed()
                await checkpoint_mgr.add_cost(cost_usd)
                if not credit_exhausted and not cost_tracker.has_sufficient_balance():
                    credit_exhausted = True
                    logger.warning("[Billing] Credits exhausted — stopping pipeline")
                    self.should_stop = True
                    stop_event.set()
            except Exception as e:
                logger.error(f"[Billing] on_cost callback error: {e}")

        if checkpoint.total_cost_usd > 0:
            cost_tracker.seed_from_checkpoint(checkpoint.total_cost_usd)

        # Detect resume
        db.refresh(version)
        existing_row_count = version.generated_count or 0
        is_resume = existing_row_count > 0

        if is_resume:
            logger.info(
                f"[Pipeline] RESUMING: {existing_row_count}/{state.num_samples} rows "
                f"already generated, checkpoint cost=${checkpoint.total_cost_usd:.4f}"
            )

        chat_history = self._load_chat_history(db, project.id)

        # Build uploaded files metadata
        project_files = (
            db.query(ProjectFile)
            .filter(
                ProjectFile.project_id == project.id,
                ProjectFile.deleted_at.is_(None),
                ProjectFile.status == "uploaded",
            )
            .all()
        )
        uploaded_files = [
            {
                "filename": f.filename,
                "content_type": f.content_type,
                "size_bytes": f.size_bytes,
            }
            for f in project_files
        ]

        if uploaded_files and uploaded_file_urls:
            uploaded_files = await self._enrich_file_metadata(
                uploaded_files, uploaded_file_urls,
            )

        # Load connectors for MCP tools
        connectors = (
            db.query(ProjectConnector)
            .filter(
                ProjectConnector.project_id == project.id,
                ProjectConnector.deleted_at.is_(None),
            )
            .all()
        )
        mcp_tools = self._build_mcp_tools(connectors)
        self._mcp_tools = mcp_tools

        logger.info(f"[Pipeline] Starting V13 pipeline with {len(chat_history)} chat messages")

        # Shared state
        generation_stats = {
            "rows_generated": existing_row_count,
            "errors": 0,
            "skipped": 0,
            "in_progress": 0,
            "total_cost": checkpoint.total_cost_usd,
        }

        # Feedback context
        feedback_context = None
        if checkpoint.recipe:
            try:
                prev_config = json.loads(checkpoint.recipe)
                feedback_msg = (
                    db.query(ChatMessage)
                    .filter(
                        ChatMessage.project_id == project.id,
                        ChatMessage.role == "user",
                        ChatMessage.created_at > version.started_at,
                    )
                    .order_by(ChatMessage.created_at.desc())
                    .first()
                )
                if feedback_msg:
                    feedback_context = {
                        "previous_config": prev_config,
                        "user_feedback": feedback_msg.content,
                    }
            except (json.JSONDecodeError, Exception):
                pass

        # Progress tracking
        version.progress_detail = {"phase": "orchestrating"}
        db.commit()

        dedup_store = DedupStore()

        # Seed DedupStore with existing rows for resume
        if is_resume:
            from dsl_worker.agents.row import _tokenize
            existing_samples = (
                db.query(Sample)
                .filter(Sample.version_id == version.id)
                .all()
            )
            for sample in existing_samples:
                row_data = sample.row or {}
                row_id = str(sample.id)
                dedup_store._submitted[row_id] = row_data
                for col_name, value in row_data.items():
                    dedup_store._token_cache[(row_id, col_name)] = _tokenize(value)
            logger.info(f"[Pipeline] Seeded DedupStore with {len(existing_samples)} existing rows")

        from dsl_worker.infra.bu_client import BUClient
        bu_client = BUClient(
            api_key=settings.browser_use_api_key,
            model="bu-mini",
            stop_event=stop_event,
        )

        apollo_client = None
        if settings.apollo_api_key:
            from dsl_worker.infra.apollo_client import ApolloClient
            apollo_client = ApolloClient(api_key=settings.apollo_api_key)

        google_maps_client = None
        if settings.google_api_key:
            from dsl_worker.infra.google_maps_client import GoogleMapsClient
            google_maps_client = GoogleMapsClient(api_key=settings.google_api_key)

        youtube_client = None
        if settings.google_api_key:
            from dsl_worker.infra.youtube_client import YouTubeClient
            youtube_client = YouTubeClient(api_key=settings.google_api_key)

        # Row saver
        row_saver = GenerationWorkerPool(
            workspace_dir=workspace_dir,
            openai_client=tracked_client,
            db_session=db,
            project_id=project.id,
            version_id=version.id,
        )

        async def save_row(row: Dict, tags: Dict = None) -> Optional[str]:
            db.refresh(version)
            if (version.generated_count or 0) >= state.num_samples:
                return None
            return await row_saver._save_row(row, tags=tags or {})

        # Resume context
        resume_context = None
        if is_resume:
            resume_context = {
                "rows_generated": existing_row_count,
                "target": state.num_samples,
                "remaining": state.num_samples - existing_row_count,
                "prior_cost_usd": checkpoint.total_cost_usd,
                "previous_sources": (checkpoint.source_stats or {}).get("sources", []),
            }

        # V13 checkpoint callback
        def on_checkpoint(orch):
            try:
                state_data = orch.export_state()
                asyncio.create_task(checkpoint_mgr.save_pipeline_state(
                    orchestrator_messages=state_data["orchestrator_conversation"]["messages"],
                    orchestrator_cost=state_data["orchestrator_conversation"]["total_cost"],
                    orchestrator_turns=state_data["orchestrator_conversation"]["total_turns"],
                    sources=[],
                    generation_stats=state_data["generation_stats"],
                    harvester_counter=0,
                    apollo_counter=0,
                    research_counter=state_data.get("web_research_counter", 0),
                ))
            except Exception as e:
                logger.warning(f"[Pipeline] Checkpoint callback error: {e}")

        # Row generation function — V13 passes Candidate objects with metadata
        _save_lock = asyncio.Lock()

        async def generate_row_fn(candidate, source_id):
            """Row generation function for V13 orchestrator."""
            rows_done = generation_stats.get("rows_generated", 0)
            if rows_done >= state.num_samples:
                return GeneratedRow(success=False, skipped=True, skip_reason="target reached"), 0.0, False

            def _row_stop():
                if stop_checker and stop_checker():
                    return True
                return generation_stats.get("rows_generated", 0) >= state.num_samples

            agent = RowGeneratorAgent(
                openai_client=tracked_client,
                model=settings.generation_model,
                workspace_dir=workspace_dir,
                chat_history=chat_history,
                dedup_store=dedup_store,
                bu_client=bu_client,
                sandbox=self._sandbox,
                stop_checker=_row_stop,
                stop_event=stop_event,
                blob_service_client=self.blob_service_client,
                project_id=project.id,
                uploaded_file_urls=uploaded_file_urls if uploaded_file_urls else None,
                uploaded_files=uploaded_files if uploaded_files else None,
                apollo_client=apollo_client,
                google_maps_client=google_maps_client,
                youtube_client=youtube_client,
                mcp_tools=mcp_tools,
                on_cost=on_cost,
            )
            try:
                # Extract V13-specific fields from Candidate metadata
                metadata = getattr(candidate, "metadata", {}) or {}
                preset_fields = metadata.get("preset_fields")
                candidate_data = metadata.get("candidate_data")
                source_context = getattr(candidate, "source_context", "") or ""

                result = await agent.generate(
                    candidate=candidate.values,
                    schema=state.columns,
                    source_context=source_context,
                    note=source_context,  # V13: note is the orchestrator's handoff briefing
                    preset_fields=preset_fields,
                    candidate_data=candidate_data,
                )

                saved = False
                if result.success and result.row:
                    async with _save_lock:
                        row_id = await save_row(
                            result.row,
                            tags={"sources": result.sources} if result.sources else {},
                        )
                        saved = row_id is not None
                        if saved:
                            generation_stats["rows_generated"] = (
                                generation_stats.get("rows_generated", 0) + 1
                            )
                elif result.skipped:
                    generation_stats["skipped"] = generation_stats.get("skipped", 0) + 1
                else:
                    generation_stats["errors"] = generation_stats.get("errors", 0) + 1

                return result, result.cost_usd, saved

            except asyncio.CancelledError:
                return GeneratedRow(success=False, skipped=True, skip_reason="stopped"), 0.0, False
            except Exception as e:
                logger.error(f"Row generation error: {e}", exc_info=True)
                return GeneratedRow(success=False, error=str(e)), 0.0, False
            finally:
                try:
                    await agent.cleanup()
                except Exception:
                    pass

        # Build the V13 orchestrator
        orchestrator = OrchestratorV13(
            chat_history=chat_history,
            columns=state.columns,
            num_samples=state.num_samples,
            openai_client=tracked_client,
            model=settings.research_model,
            workspace_dir=workspace_dir,
            generation_stats=generation_stats,
            dedup_store=dedup_store,
            save_row=save_row,
            generate_row_fn=generate_row_fn,
            uploaded_files=uploaded_files if uploaded_files else None,
            bu_client=bu_client,
            sandbox=self._sandbox,
            stop_checker=stop_checker,
            stop_event=stop_event,
            blob_service_client=self.blob_service_client,
            project_id=project.id,
            on_tool_call=None,
            on_cost=on_cost,
            on_checkpoint=on_checkpoint,
            uploaded_file_urls=uploaded_file_urls if uploaded_file_urls else None,
            mcp_tools=mcp_tools,
            apollo_client=apollo_client,
            google_maps_client=google_maps_client,
            youtube_client=youtube_client,
            feedback_context=feedback_context,
            resume_context=resume_context,
        )

        # Restore state from checkpoint if resuming
        if is_resume and checkpoint.has_v11_state:
            orchestrator.restore_state({
                "orchestrator_conversation": checkpoint.orchestrator_conversation,
                "generation_stats": checkpoint.generation_stats,
                "web_research_counter": checkpoint.research_counter or 0,
                "bu_extract_counter": 0,
                "current_file": None,
            })

        try:
            result = await orchestrator.run()
            logger.info(
                f"[Pipeline] V13 orchestrator finished: "
                f"cost=${orchestrator.cost_usd:.4f}, turns={result.turns_taken}"
            )
        finally:
            try:
                state_data = orchestrator.export_state()
                await checkpoint_mgr.save_pipeline_state(
                    orchestrator_messages=state_data["orchestrator_conversation"]["messages"],
                    orchestrator_cost=state_data["orchestrator_conversation"]["total_cost"],
                    orchestrator_turns=state_data["orchestrator_conversation"]["total_turns"],
                    sources=[],
                    generation_stats=state_data["generation_stats"],
                    harvester_counter=0,
                    apollo_counter=0,
                    research_counter=state_data.get("web_research_counter", 0),
                )
                await checkpoint_mgr.force_save()
            except Exception as e:
                logger.warning(f"[Pipeline] Final checkpoint save error: {e}")

            await orchestrator.cleanup()
            await asyncio.sleep(0.5)
            await bu_client.close()
            if apollo_client:
                await apollo_client.close()

        # === COMPLETE — Check results ===
        await checkpoint_mgr.set_phase("completed")

        state.refresh()
        if state.paused:
            await checkpoint_mgr.force_save()
            self._handle_pause(db, project, version, cost_tracker)
            return True

        can_continue, stop_reason = cost_tracker.check_balance_and_charge()
        if not can_continue:
            await checkpoint_mgr.force_save()
            self._handle_force_stop(db, project, version, cost_tracker, stop_reason)
            return False

        db.refresh(version)
        generated = version.generated_count or 0
        if generated > 0:
            if generated < state.num_samples:
                logger.warning(
                    f"[Pipeline] Completed with {generated}/{state.num_samples} rows"
                )
            await checkpoint_mgr.force_save()
            self._handle_completion(db, project, version, cost_tracker)
            return True

        await checkpoint_mgr.force_save()
        self._handle_force_stop(db, project, version, cost_tracker, "No rows generated")
        return False

    def _load_chat_history(self, db: Session, project_id: UUID) -> List[Dict[str, str]]:
        """Load chat messages and format for the orchestrator.

        Includes applied_changes data (plan, questions, columns) so the
        orchestrator has full context from the consultation chat.
        """
        messages = (
            db.query(ChatMessage)
            .filter(ChatMessage.project_id == project_id)
            .order_by(ChatMessage.created_at.asc())
            .all()
        )

        history = []
        for msg in messages:
            if msg.role not in ("user", "assistant"):
                continue

            content = msg.content or ""

            # Append structured data from applied_changes so the orchestrator
            # sees the plan, questions, and column definitions from the chat.
            if msg.applied_changes and isinstance(msg.applied_changes, dict):
                changes = msg.applied_changes.get("changes", {})

                if "plan" in changes:
                    plan = changes["plan"]
                    if isinstance(plan, dict):
                        title = plan.get("title", "")
                        bullets = plan.get("bullets", [])
                        overview = plan.get("overview", "")
                        if title and bullets:
                            plan_text = f"{title}\n" + "\n".join(f"- {b}" for b in bullets)
                            content += f"\n\n<plan>\n{plan_text}\n</plan>"
                        elif overview:
                            content += f"\n\n<plan>\n{overview}\n</plan>"
                    elif isinstance(plan, str):
                        content += f"\n\n<plan>\n{plan}\n</plan>"

                if "columns" in changes:
                    cols = changes["columns"]
                    if isinstance(cols, list) and cols:
                        content += "\n\n<columns>\n" + json.dumps(cols, indent=2) + "\n</columns>"

                if "questions" in changes:
                    qs = changes["questions"]
                    if isinstance(qs, list):
                        q_blocks = []
                        for q in qs:
                            if isinstance(q, dict):
                                question_text = q.get("question", q.get("label", "?"))
                                q_type = q.get("type", "single_choice")
                                options = q.get("options", [])
                                block = question_text
                                if q_type == "multi_choice":
                                    block += " (pick all that apply)"
                                if options:
                                    block += "\n" + "\n".join(
                                        f"  - {opt}" for opt in options
                                    )
                                q_blocks.append(block)
                            elif isinstance(q, str):
                                q_blocks.append(q)
                        if q_blocks:
                            content += "\n\n<questions>\n" + "\n\n".join(q_blocks) + "\n</questions>"

            history.append({
                "role": msg.role,
                "content": content,
            })

        return history

    async def _enrich_file_metadata(
        self,
        uploaded_files: List[Dict],
        uploaded_file_urls: Dict[str, str],
    ) -> List[Dict]:
        """Inspect uploaded files and add metadata (columns, row count, preview).

        Creates a temporary sandbox session, uploads the files, runs a quick
        inspection script, and destroys the session. Takes 1-3 seconds.
        """
        from dsl_worker.infra.sandbox import SandboxSession

        inspectable = {
            f["filename"] for f in uploaded_files
            if f.get("content_type", "") in (
                "text/csv", "application/json", "application/vnd.ms-excel",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ) or f.get("filename", "").lower().endswith((".csv", ".json", ".jsonl", ".xlsx", ".xls", ".tsv"))
        }

        if not inspectable:
            return uploaded_files

        try:
            session_client = await self._sandbox.create_session()
            session = SandboxSession(session_client, self._sandbox)

            # Upload files
            urls_to_fetch = {k: v for k, v in uploaded_file_urls.items() if k in inspectable}
            await session.upload_workspace(self._workspace_dir, urls_to_fetch)

            # Inspection script
            script = '''
import os, json, csv

results = {}
upload_dir = "/workspace/uploads"
for fname in os.listdir(upload_dir):
    path = os.path.join(upload_dir, fname)
    if not os.path.isfile(path):
        continue
    info = {}
    try:
        lower = fname.lower()
        if lower.endswith(".csv") or lower.endswith(".tsv"):
            delim = "\\t" if lower.endswith(".tsv") else ","
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                reader = csv.DictReader(f, delimiter=delim)
                cols = reader.fieldnames or []
                rows = sum(1 for _ in reader)
            info = {"type": "csv", "columns": cols, "row_count": rows}
            # Preview first 3 rows
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                reader = csv.DictReader(f, delimiter=delim)
                preview = []
                for i, row in enumerate(reader):
                    if i >= 3:
                        break
                    preview.append(dict(row))
                info["preview"] = preview
        elif lower.endswith(".json"):
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                data = json.load(f)
            if isinstance(data, list):
                info = {"type": "json_array", "item_count": len(data)}
                if data and isinstance(data[0], dict):
                    info["keys"] = list(data[0].keys())
                    info["preview"] = data[:3]
            elif isinstance(data, dict):
                info = {"type": "json_object", "keys": list(data.keys())[:20]}
        elif lower.endswith(".jsonl"):
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                lines = f.readlines()
            info = {"type": "jsonl", "line_count": len(lines)}
            if lines:
                first = json.loads(lines[0])
                if isinstance(first, dict):
                    info["keys"] = list(first.keys())
                    info["preview"] = [json.loads(l) for l in lines[:3]]
        elif lower.endswith((".xlsx", ".xls")):
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheets = wb.sheetnames
            ws = wb[sheets[0]]
            rows_list = list(ws.iter_rows(max_row=4, values_only=True))
            wb.close()
            cols = [str(c) if c else "" for c in rows_list[0]] if rows_list else []
            info = {"type": "xlsx", "sheets": sheets, "columns": cols, "row_count": ws.max_row - 1 if ws.max_row else 0}
            if len(rows_list) > 1:
                info["preview"] = [dict(zip(cols, [str(v) if v else "" for v in r])) for r in rows_list[1:4]]
    except Exception as e:
        info = {"error": str(e)}
    results[fname] = info

print(json.dumps(results))
'''
            result = await session.execute(script, timeout=30)
            await session.close()

            if result.success and result.stdout.strip():
                try:
                    file_info = json.loads(result.stdout.strip())
                except json.JSONDecodeError:
                    logger.warning(f"[Pipeline] File inspection output not JSON: {result.stdout[:200]}")
                    return uploaded_files

                enriched = []
                for f in uploaded_files:
                    entry = dict(f)
                    info = file_info.get(f["filename"])
                    if info and "error" not in info:
                        entry["inspection"] = info
                        logger.info(
                            f"[Pipeline] Inspected {f['filename']}: "
                            f"type={info.get('type')}, "
                            f"rows={info.get('row_count', info.get('item_count', info.get('line_count', '?')))}, "
                            f"cols={len(info.get('columns', info.get('keys', [])))}"
                        )
                    enriched.append(entry)
                return enriched
            else:
                logger.warning(f"[Pipeline] File inspection failed: {result.stderr[:200] if result.stderr else 'no output'}")
                return uploaded_files

        except Exception as e:
            logger.warning(f"[Pipeline] File inspection error: {e}")
            return uploaded_files

    def _restore_downloads_from_blob(self, project_id, workspace_dir: Path) -> None:
        """Restore browser-downloaded files from blob into workspace/downloads/.

        Files are uploaded to blob during execution by ResearchTools._upload_download_to_blob().
        On resume, we restore them so the sandbox and agents have access again.
        """
        prefix = f"projects/{project_id}/downloads/"
        downloads_dir = workspace_dir / "downloads"
        try:
            container_client = self.blob_service_client.get_container_client(
                settings.azure_storage_container_name
            )
            blobs = list(container_client.list_blobs(name_starts_with=prefix))
            if not blobs:
                return

            for blob in blobs:
                filename = blob.name[len(prefix):]
                if not filename:
                    continue
                local_path = downloads_dir / filename
                try:
                    blob_client = container_client.get_blob_client(blob.name)
                    data = blob_client.download_blob().readall()
                    local_path.write_bytes(data)
                    logger.info(f"[Pipeline] Restored download: {filename} ({len(data)} bytes)")
                except Exception as e:
                    logger.warning(f"[Pipeline] Failed to restore {filename}: {e}")

            logger.info(f"[Pipeline] Restored {len(blobs)} browser downloads from blob")
        except Exception as e:
            logger.debug(f"[Pipeline] No downloads to restore: {e}")

    def _generate_file_urls(self, db: Session, project_id) -> Dict[str, str]:
        """Generate short-lived SAS URLs for uploaded files (no local download needed).

        Returns dict of filename -> SAS URL. The sandbox service will fetch
        these URLs directly into the session workspace.
        """
        project_files = (
            db.query(ProjectFile)
            .filter(
                ProjectFile.project_id == project_id,
                ProjectFile.deleted_at.is_(None),
                ProjectFile.status == "uploaded",
            )
            .all()
        )
        if not project_files:
            logger.info("[Pipeline] No uploaded files")
            return {}

        urls: Dict[str, str] = {}
        container = settings.azure_storage_container_name

        for f in project_files:
            filename = f.filename
            blob_path = f.blob_path
            if not filename or not blob_path:
                continue

            try:
                sas_token = generate_blob_sas(
                    account_name=settings.azure_storage_account_name,
                    account_key=settings.azure_storage_account_key,
                    container_name=container,
                    blob_name=blob_path,
                    permission=BlobSasPermissions(read=True),
                    expiry=datetime.now(timezone.utc) + timedelta(hours=1),
                )
                blob_url = (
                    f"https://{settings.azure_storage_account_name}.blob.core.windows.net"
                    f"/{container}/{blob_path}?{sas_token}"
                )
                urls[filename] = blob_url
                logger.info(f"[Pipeline] Generated SAS URL for: {filename}")
            except Exception as e:
                logger.error(f"[Pipeline] Failed to generate SAS URL for {filename}: {e}")

        return urls

    def _build_mcp_tools(self, connectors) -> list:
        """Build MCP tool definitions from project connectors."""
        from dsl_api.crypto import decrypt_secret

        mcp_tools = []
        for conn in connectors:
            tool_def = {"type": "mcp", "server_label": conn.server_label}

            if conn.server_url:
                tool_def["server_url"] = conn.server_url
            if conn.connector_id:
                tool_def["connector_id"] = conn.connector_id
            if conn.server_description:
                tool_def["server_description"] = conn.server_description
            if conn.allowed_tools is not None:
                tool_def["allowed_tools"] = conn.allowed_tools

            # Decrypt auth
            if conn.authorization_encrypted:
                tool_def["authorization"] = decrypt_secret(conn.authorization_encrypted)
            if conn.headers_encrypted:
                import json as _json
                tool_def["headers"] = _json.loads(decrypt_secret(conn.headers_encrypted))

            tool_def["require_approval"] = "never"
            mcp_tools.append(tool_def)

        return mcp_tools

    async def _cleanup(self):
        """Cleanup resources."""
        if self._sandbox:
            try:
                await self._sandbox.__aexit__(None, None, None)
            except Exception as e:
                logger.warning(f"Sandbox cleanup error: {e}")
            self._sandbox = None

        # Clean up workspace directory
        workspace_dir = getattr(self, "_workspace_dir", None)
        if workspace_dir and workspace_dir.exists():
            try:
                shutil.rmtree(workspace_dir)
                logger.info(f"Cleaned up workspace: {workspace_dir}")
            except Exception as e:
                logger.warning(f"Workspace cleanup error: {e}")
            self._workspace_dir = None

    def _emit_event(
        self,
        db: Session,
        project: Project,
        version: ProjectVersion,
        event_type: str,
        message: str,
        details: dict = None
    ) -> None:
        """Emit project event."""
        event = ProjectEvent(
            project_id=project.id,
            version_id=version.id,
            event_type=event_type,
            message=message,
            details=details or {}
        )
        db.add(event)
        db.commit()

    def _handle_pause(
        self,
        db: Session,
        project: Project,
        version: ProjectVersion,
        cost_tracker: Optional[CostTracker],
        message: str = "Worker paused"
    ) -> None:
        """Handle pause."""
        logger.info(f"Pausing version {version.id}")
        db.refresh(version)

        if cost_tracker:
            cost_tracker.charge_remaining()

        version.status = "paused"

        details = {"paused_at": datetime.now(timezone.utc).isoformat()}
        if cost_tracker:
            summary = cost_tracker.get_summary()
            details["total_cost_cents"] = summary["total_costs_cents"]

        self._emit_event(db, project, version, "paused", message, details)
        db.commit()

    def _handle_completion(
        self,
        db: Session,
        project: Project,
        version: ProjectVersion,
        cost_tracker: CostTracker
    ) -> None:
        """Handle successful completion."""
        logger.info(f"Version {version.id} completed")
        db.refresh(version)

        cost_tracker.charge_remaining()

        version.status = "succeeded"
        version.finished_at = datetime.now(timezone.utc)
        version.progress_detail = None

        summary = cost_tracker.get_summary()
        self._emit_event(
            db, project, version, "completed",
            "Generation complete",
            {
                "completed_at": datetime.now(timezone.utc).isoformat(),
                "generated_count": version.generated_count,
                "total_cost_cents": summary["total_costs_cents"],
            }
        )

        db.commit()
        logger.info(f"Completed: {version.generated_count} samples")

    def _handle_force_stop(
        self,
        db: Session,
        project: Project,
        version: ProjectVersion,
        cost_tracker: CostTracker,
        reason: str
    ) -> None:
        """Handle force-stop. Credit exhaustion → paused (resumable), other → failed."""
        logger.warning(f"Force-stopping: {reason}")
        db.refresh(version)

        cost_tracker.charge_remaining()

        is_credit_stop = reason in ("insufficient_balance", "credit_exhausted")
        generated = version.generated_count or 0

        if is_credit_stop and generated > 0:
            # Credit exhaustion with partial progress → paused (user can buy credits & resume)
            version.status = "paused"
            version.error = reason
            self._emit_event(
                db, project, version, "paused",
                f"Paused: {reason} ({generated} rows generated)",
                {"reason": reason, "generated_count": generated}
            )
        else:
            # Real failure or no progress
            version.status = "failed"
            version.error = reason
            version.finished_at = datetime.now(timezone.utc)
            version.progress_detail = None
            self._emit_event(
                db, project, version, "failed",
                reason,
                {"reason": reason}
            )

        db.commit()
