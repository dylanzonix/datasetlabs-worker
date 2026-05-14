"""file source — parse uploaded CSV/XLSX or files written by code_exec.

CSV/XLSX get auto-detected and parsed; rows = file rows, columns inferred
from headers (with light type sniffing). Other file types return an error
with a hint to use `code_exec` to transform first.

Predictable in the sense that headers map directly to columns — no agent-side
mapping step needed. (Agent can still rename/retype columns later via
`column_map_set` if desired.)
"""

from __future__ import annotations

import csv
import io
import logging
import os
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

from dsl_worker.sources_v2.base import FetchResult, SourceAdapter, register


log = logging.getLogger(__name__)


_INT_RE = re.compile(r"^-?\d+$")
_FLOAT_RE = re.compile(r"^-?\d+\.\d+$")
_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}")
_URL_RE = re.compile(r"^https?://")
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _sniff_type(samples: List[str]) -> str:
    """Infer a column's type from a handful of sample values. Conservative."""
    non_empty = [s for s in samples if s]
    if not non_empty:
        return "text"
    if all(_URL_RE.match(s) for s in non_empty):
        return "url"
    if all(_EMAIL_RE.match(s) for s in non_empty):
        return "email"
    if all(_DATE_RE.match(s) for s in non_empty):
        return "date"
    if all(_FLOAT_RE.match(s) or _INT_RE.match(s) for s in non_empty):
        return "number"
    if all(s.lower() in {"true", "false", "yes", "no", "1", "0"} for s in non_empty):
        return "bool"
    return "text"


class FileAdapter(SourceAdapter):
    name = "file"
    predictable = True  # CSV headers ARE the schema

    def validate_query_params(self, query_params: Dict[str, Any]) -> Optional[str]:
        if "file_id" not in query_params:
            return "file source requires `file_id`"
        return None

    async def fetch(
        self,
        query_params: Dict[str, Any],
        n: int,
        prior_cursor: Optional[Dict[str, Any]] = None,
    ) -> FetchResult:
        file_id = query_params["file_id"]
        path = Path(file_id)
        # If not on disk, try the candidate store (files written by
        # code_exec are uploaded there).
        if not path.exists():
            project_id = query_params.get("_project_id")
            if project_id:
                try:
                    from dsl_worker.chat_api.candidates import read_candidates_bytes
                    blob_bytes = read_candidates_bytes(project_id, file_id)
                    import tempfile
                    suffix = Path(file_id).suffix or ".csv"
                    tmp = tempfile.NamedTemporaryFile(
                        delete=False, suffix=suffix, mode="wb",
                    )
                    tmp.write(blob_bytes)
                    tmp.close()
                    path = Path(tmp.name)
                except Exception as e:
                    log.info("file source: candidate lookup for %s failed: %s", file_id, e)
                    return FetchResult(
                        rows=[], schema=[], cost_credits=0.0, exhausted=True,
                    )
            else:
                return FetchResult(
                    rows=[], schema=[], cost_credits=0.0, exhausted=True,
                )

        rows: List[Dict[str, Any]] = []
        if path.suffix.lower() == ".csv":
            with path.open("r", encoding="utf-8", errors="replace") as f:
                reader = csv.DictReader(f)
                offset = int((prior_cursor or {}).get("offset", 0))
                for i, row in enumerate(reader):
                    if i < offset:
                        continue
                    if len(rows) >= n:
                        break
                    rows.append({k: v for k, v in row.items() if k})
        else:
            # XLSX support — try openpyxl if installed. Otherwise error.
            try:
                import openpyxl
                wb = openpyxl.load_workbook(path, read_only=True)
                ws = wb.active
                headers = [str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
                offset = int((prior_cursor or {}).get("offset", 0))
                for i, row_cells in enumerate(ws.iter_rows(min_row=2)):
                    if i < offset:
                        continue
                    if len(rows) >= n:
                        break
                    rows.append({h: (c.value if c.value is not None else "") for h, c in zip(headers, row_cells)})
            except ImportError:
                return FetchResult(
                    rows=[],
                    schema=[],
                    cost_credits=0.0,
                    exhausted=True,
                )
            except Exception as e:
                log.exception("file adapter xlsx parse failed: %s", e)
                return FetchResult(rows=[], schema=[], cost_credits=0.0, exhausted=True)

        # Sniff types from first 20 rows to populate default_columns dynamically.
        if rows:
            schema_keys = list({k for r in rows for k in r.keys()})
            sniffed: List[Dict[str, str]] = []
            for k in schema_keys:
                samples = [str(r.get(k, "") or "") for r in rows[:20]]
                sniffed.append({"source_field": k, "column_name": k, "type": _sniff_type(samples)})
            self.default_columns = sniffed

        return FetchResult(
            rows=rows,
            schema=list(rows[0].keys()) if rows else [],
            cost_credits=0.0,
            exhausted=len(rows) < n,
            cursor={"offset": int((prior_cursor or {}).get("offset", 0)) + len(rows)},
        )


register(FileAdapter())
